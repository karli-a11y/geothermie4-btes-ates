#!/usr/bin/env python3
"""
Berechnet aus der Berechnungshilfe-Tabelle (Solarthermie SS-2025) und einem
monatlichen Wärme­bedarf eine 12-Werte-Liste `monthly_power_W`, die direkt in
den `CONFIG["cycles"]["monthly_power_W"]`-Block der BTES- oder ATES-Skripte
eingesetzt werden kann.

Konzept
-------
Eingang  : - solarer Monats­ertrag q_sol(m) [kWh/m²/Monat]   (aus xlsx)
           - Kollektor­fläche A_koll [m²]
           - Wärme­bedarf Q_bed(m) [kWh/Monat]
Bilanz   : ΔQ(m) = A_koll · q_sol(m) − Q_bed(m)   [kWh/Monat]
           + Überschuss → in den Speicher laden
           − Defizit    → aus dem Speicher fördern
Leistung : P(m) = ΔQ(m) · 3600 · 1000 / ( days(m) · 86400 )   [W]

Hinweis zur Auslegung: damit der Speicher nicht überdimensioniert wird,
sollte die Jahres­bilanz Σ ΔQ(m) ≈ 0 gewählt werden. Das heißt A_koll wird
so gewählt, dass A_koll · Σ q_sol(m) ≈ Σ Q_bed(m).

Verwendung
----------
    python solar_to_monthly.py
        → Demo mit Default-Bedarf, druckt die monthly_power_W-Liste.

Als Modul:
    from solar_to_monthly import monthly_power_W
    P = monthly_power_W(A_koll_m2=30.0, demand_kWh_per_month=[...12 Werte...])
    CONFIG["cycles"]["monthly_power_W"] = P
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:                    # pragma: no cover
    raise SystemExit("openpyxl fehlt – installieren mit:  pip install openpyxl") from e


XLSX_DEFAULT  = str(Path(__file__).parent / "data" / "Solarthermie_Berechnungshilfe.xlsx")
SHEET_DEFAULT = "Vakuum-Röhrenkollektor"   # alternativ: "Flachkollektor"
BETA_DEFAULT  = 40                         # Kollektor-Anstellwinkel [°]

DAYS_PER_MONTH = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
MONTHS         = ["Jan", "Feb", "Mrz", "Apr", "Mai", "Jun",
                  "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]


def solar_monthly_yield(xlsx_path: str | Path = XLSX_DEFAULT,
                        sheet: str = SHEET_DEFAULT,
                        beta_deg: float = BETA_DEFAULT) -> list[float]:
    """Liest die 12 Monats­erträge q_sol [kWh/m²/Monat] für gewählten Anstellwinkel."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' nicht in {xlsx_path}; verfügbar: {wb.sheetnames}")
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=True):
        if len(row) >= 14 and isinstance(row[1], (int, float)) and float(row[1]) == float(beta_deg):
            vals = list(row[2:14])
            if all(isinstance(v, (int, float)) for v in vals):
                return [float(v) for v in vals]
    raise ValueError(f"Anstellwinkel β = {beta_deg}° nicht in Sheet '{sheet}' gefunden.")


def monthly_power_W(A_koll_m2: float,
                    demand_kWh_per_month: list[float] | tuple[float, ...],
                    *,
                    xlsx_path: str | Path = XLSX_DEFAULT,
                    sheet: str = SHEET_DEFAULT,
                    beta_deg: float = BETA_DEFAULT) -> list[float]:
    """
    Liefert die 12-Werte-Liste `monthly_power_W` [W].

    A_koll_m2            – Kollektor­fläche [m²]
    demand_kWh_per_month – Wärme­bedarf je Monat (Liste mit 12 Werten) [kWh]
    """
    if len(demand_kWh_per_month) != 12:
        raise ValueError("demand_kWh_per_month muss exakt 12 Werte enthalten.")
    q_sol = solar_monthly_yield(xlsx_path, sheet, beta_deg)
    P = []
    for m in range(12):
        delta_Q_kWh = A_koll_m2 * q_sol[m] - float(demand_kWh_per_month[m])
        seconds = DAYS_PER_MONTH[m] * 86400.0
        P.append(delta_Q_kWh * 3.6e6 / seconds)         # kWh → J, /s = W
    return P


def annual_balance_kWh(A_koll_m2: float,
                        demand_kWh_per_month: list[float] | tuple[float, ...],
                        *,
                        xlsx_path: str | Path = XLSX_DEFAULT,
                        sheet: str = SHEET_DEFAULT,
                        beta_deg: float = BETA_DEFAULT) -> float:
    """Σ_m (A · q_sol − Q_bed) – Vorzeichen-Indikator für Auslegung."""
    q_sol = solar_monthly_yield(xlsx_path, sheet, beta_deg)
    return A_koll_m2 * sum(q_sol) - float(sum(demand_kWh_per_month))


def sizing_A_koll_for_balance(demand_kWh_per_month, *,
                              xlsx_path: str | Path = XLSX_DEFAULT,
                              sheet: str = SHEET_DEFAULT,
                              beta_deg: float = BETA_DEFAULT) -> float:
    """Kollektor­fläche [m²] so, dass Jahres­bilanz exakt 0 ergibt."""
    q_sol_year = sum(solar_monthly_yield(xlsx_path, sheet, beta_deg))
    return float(sum(demand_kWh_per_month)) / q_sol_year


# Typischer normierter Heiz­bedarfs­verlauf (Anteil am Jahres­bedarf), Mitteleuropa:
DEMAND_PROFILE_FRAC = [0.175, 0.145, 0.120, 0.080, 0.045, 0.020,
                       0.010, 0.015, 0.040, 0.095, 0.125, 0.130]


def demand_from_annual(Q_annual_kWh: float,
                       profile: list[float] = DEMAND_PROFILE_FRAC) -> list[float]:
    """Skaliert das normierte Lastprofil (Σ = 1) auf den vorgegebenen Jahres­bedarf."""
    s = sum(profile)
    return [Q_annual_kWh * (p / s) for p in profile]


def _print_table(q_sol, demand, P):
    print(f"{'Monat':<5} {'q_sol':>10} {'A·q_sol':>10} {'Q_bed':>10} {'ΔQ':>10} {'P':>12}")
    print(f"{'':<5} {'[kWh/m²]':>10} {'[kWh]':>10} {'[kWh]':>10} {'[kWh]':>10} {'[W]':>12}")
    for m in range(12):
        print(f"{MONTHS[m]:<5} {q_sol[m]:>10.1f} {q_sol[m]*A_KOLL_DEMO:>10.0f}"
              f" {demand[m]:>10.0f} {q_sol[m]*A_KOLL_DEMO - demand[m]:>10.0f}"
              f" {P[m]:>12.1f}")


# ----------------------------------------------------------------------
# Demo
# ----------------------------------------------------------------------
A_KOLL_DEMO       = 30.0           # m² – Kollektor­fläche
Q_ANNUAL_DEMO     = 25_000.0       # kWh/a – Jahres­wärmebedarf


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    print(f"Eingaben:  A_koll = {A_KOLL_DEMO} m²,  Q_jährlich = {Q_ANNUAL_DEMO:.0f} kWh")
    print(f"Sheet      {SHEET_DEFAULT!r},  β = {BETA_DEFAULT}°")

    q_sol  = solar_monthly_yield(beta_deg=BETA_DEFAULT)
    demand = demand_from_annual(Q_ANNUAL_DEMO)
    P      = monthly_power_W(A_KOLL_DEMO, demand)
    _print_table(q_sol, demand, P)

    bal = annual_balance_kWh(A_KOLL_DEMO, demand)
    A_balanced = sizing_A_koll_for_balance(demand)
    print()
    print(f"Jahres­bilanz Σ ΔQ:  {bal:+.0f} kWh"
          f"   →  {'Überschuss (Speicher überdimensioniert)' if bal > 0 else 'Defizit'}")
    print(f"Empfohlene Fläche für  Σ ΔQ = 0:  A_koll ≈ {A_balanced:.1f} m²")

    print()
    print("# In OGS-CONFIG einsetzen:")
    print("CONFIG['cycles']['monthly_power_W'] = [")
    for p in P:
        print(f"    {p:.1f},")
    print("]")
